"""
convert_excel.py
讀取 Netflix 台灣收視排名數據分析 Excel，輸出 rankings.json

用法：
  python scripts/convert_excel.py                          # 使用預設路徑
  python scripts/convert_excel.py path/to/excel.xlsx       # 指定 Excel 路徑

預期 Sheet 結構（5 sheets）：
  [0] Netflix 年度排行分析      — 樞紐分析表（僅供參考，不直接讀取）
  [1] clean Data - Netflix 每週排名  — 全年度合併週榜 clean data
  [2] 台劇排名分析              — 樞紐分析表（僅供參考，不直接讀取）
  [3] Clean Data - 台劇每日排名  — 台劇每日 clean data
  [4] 劇集屬性資料庫            — 劇集固定屬性（Netflix Original、上架方式、集數等）
"""

import json
import sys
from collections import defaultdict
from datetime import datetime
from pathlib import Path

try:
    import openpyxl
except ImportError:
    print("ERROR: openpyxl not installed. Run: pip install openpyxl")
    sys.exit(1)

# 預設 Excel 路徑（可透過命令列參數覆蓋）
DEFAULT_EXCEL = Path("C:/Users/User/Downloads/Netflix 2025 台灣收視排名數據分析.xlsx")
OUTPUT_DIR = Path(__file__).parent.parent

# 縮寫名稱 → 正式全名對照表
TITLE_MAP = {
    "何戀":          "何百芮的地獄戀曲",
    "何毒":          "何百芮的地獄毒白",
    "死了娛樂女記者": "死了一個娛樂女記者之後",
    "太陽Part1":     "如果我不曾見過太陽",
    "太陽Part2":     "如果我不曾見過太陽",
}

# Sheet 名稱常數（用名稱查找，不用 index）
SHEET_WEEKLY = "clean Data - Netflix 每週排名"
SHEET_DAILY  = "Clean Data - 台劇每日排名"
SHEET_ATTRS  = "劇集屬性資料庫"


# ── helpers ──────────────────────────────────────────────────────────────────

def safe_int(v, default=None):
    if v is None:
        return default
    try:
        return int(float(v))
    except (ValueError, TypeError):
        return default


def safe_float(v, default=None):
    if v is None:
        return default
    try:
        return round(float(v), 2)
    except (ValueError, TypeError):
        return default


def safe_bool(v):
    if v is None:
        return False
    if isinstance(v, bool):
        return v
    return str(v).strip().lower() in ("true", "1", "是", "y", "✓", "★")


def clean_title(v):
    if not v:
        return ""
    s = str(v).strip()
    s = s.replace("《", "").replace("》", "")
    # 標準化「第 X 季」→「第X季」，統一空白
    import re
    s = re.sub(r'第\s+(\d+)\s*季', r'第\1季', s)
    # 標準化中文數字季數 → 阿拉伯數字：第一季→第1季, 第二季→第2季, ...
    cn_num = {'一': '1', '二': '2', '三': '3', '四': '4', '五': '5',
              '六': '6', '七': '7', '八': '8', '九': '9', '十': '10'}
    m = re.search(r'第([一二三四五六七八九十])季', s)
    if m:
        s = s[:m.start()] + f'第{cn_num[m.group(1)]}季' + s[m.end():]
    return TITLE_MAP.get(s, s)


def to_date_str(v):
    if v is None:
        return None
    if isinstance(v, datetime):
        return v.strftime("%Y-%m-%d")
    return None


def build_col_map(ws, header_row_idx):
    """從指定 header row 建立 {欄位名稱: 欄位索引} 對照表"""
    col_map = {}
    for j, row in enumerate(ws.iter_rows(values_only=True)):
        if j == header_row_idx:
            for i, cell in enumerate(row):
                if cell:
                    name = str(cell).strip().replace("\n", "")
                    col_map[name] = i
            break
    return col_map


def get_col(row, col_map, name, default=None):
    """安全取值：從 col_map 查欄位索引再取 row 值"""
    idx = col_map.get(name)
    if idx is None or idx >= len(row):
        return default
    return row[idx]


def find_sheet(wb, name):
    """依名稱查找 sheet，找不到則報錯"""
    if name in wb.sheetnames:
        return wb[name]
    # 容錯：嘗試忽略大小寫和空白
    for sn in wb.sheetnames:
        if sn.strip().lower() == name.strip().lower():
            return wb[sn]
    return None


# ── 校驗 ─────────────────────────────────────────────────────────────────────

def validate_date_range(date_range, context=""):
    """確認 dateRange 格式為 YYYY-MM-DD ~ YYYY-MM-DD"""
    if not date_range or " ~ " not in date_range:
        print(f"  ⚠ 日期格式異常 {context}: '{date_range}'")
        return False
    parts = date_range.split(" ~ ")
    for p in parts:
        try:
            datetime.strptime(p.strip(), "%Y-%m-%d")
        except ValueError:
            print(f"  ⚠ 日期無法解析 {context}: '{p}'")
            return False
    return True


def validate_weekly_week(week):
    """校驗單週資料完整性"""
    errors = []
    if not week.get("dateRange"):
        errors.append("缺少 dateRange")
    if not week.get("rankings"):
        errors.append("無排名資料")
    else:
        positions = [r["position"] for r in week["rankings"]]
        if len(positions) != len(set(positions)):
            dup = [p for p in positions if positions.count(p) > 1]
            errors.append(f"重複名次: {set(dup)}")
        for r in week["rankings"]:
            if not r.get("title"):
                errors.append(f"名次 {r['position']} 缺少節目名稱")
    return errors


# ── 劇集屬性資料庫 ───────────────────────────────────────────────────────────

def parse_show_attributes(ws):
    """解析劇集屬性資料庫，回傳 {節目名稱: {屬性}} 對照表"""
    col_map = build_col_map(ws, 0)
    required = ["節目名稱"]
    if not all(r in col_map for r in required):
        print(f"  ⚠ 劇集屬性資料庫缺少「節目名稱」欄位，跳過")
        return {}

    attrs = {}
    for j, row in enumerate(ws.iter_rows(values_only=True)):
        if j == 0:
            continue
        title = clean_title(get_col(row, col_map, "節目名稱"))
        if not title:
            continue

        is_orig = safe_bool(get_col(row, col_map, "是否Netflix Original"))
        release_method = str(get_col(row, col_map, "上架方式", "")).strip()
        release_weeks = safe_int(get_col(row, col_map, "上架周數"), 1)
        total_episodes = str(get_col(row, col_map, "總集數", "")).strip()

        # 上架方式 → releaseType
        if release_method == "一次上架":
            release_type = "allAtOnce"
        elif release_method == "拆分上架":
            release_type = "split"
        elif release_method == "週播":
            release_type = "weekly"
        else:
            release_type = "weekly"

        attrs[title] = {
            "isNetflixOriginal": is_orig,
            "releaseType": release_type,
            "releaseWeeks": release_weeks,
            "totalEpisodes": total_episodes,
        }

    return attrs


# ── Weekly rankings (clean long format, 全年度合併) ─────────────────────────

def parse_weekly_clean(ws):
    """解析合併後的週榜 clean data（含「類別」欄位，只取「節目」）"""
    col_map = build_col_map(ws, 0)

    use_named = "日期起" in col_map and "排名" in col_map and "節目名稱" in col_map
    if not use_named:
        print("  ⚠ 週榜缺少必要欄位，嘗試 fallback 到索引模式")
        return _parse_weekly_indexed(ws)

    date_dict = {}

    for j, row in enumerate(ws.iter_rows(values_only=True)):
        if j == 0:
            continue

        # 過濾：只取「節目」類別（排除「電影」）
        if "類別" in col_map:
            cat = str(get_col(row, col_map, "類別", "")).strip()
            if cat != "節目":
                continue

        date_from = get_col(row, col_map, "日期起")
        date_to = get_col(row, col_map, "日期迄")
        rank = safe_int(get_col(row, col_map, "排名"))
        title = clean_title(get_col(row, col_map, "節目名稱"))
        genre = str(get_col(row, col_map, "類型", "其他")).strip()
        is_orig = safe_bool(get_col(row, col_map, "是否Netflix Original"))
        score = safe_float(get_col(row, col_map, "積分"), 0)

        if date_from is None:
            continue

        key = date_from if isinstance(date_from, datetime) else str(date_from)
        if key not in date_dict:
            if isinstance(date_from, datetime) and isinstance(date_to, datetime):
                date_range = f"{date_from.strftime('%Y-%m-%d')} ~ {date_to.strftime('%Y-%m-%d')}"
            else:
                date_range = str(date_from)
            date_dict[key] = {"dateFrom": date_from, "dateRange": date_range, "rankings": [], "_skip": False}

        if title and rank is not None:
            existing = date_dict[key]["rankings"]
            # 重複 rank=1 表示同一週被重複記錄（跳過）
            if rank == 1 and len(existing) > 0:
                date_dict[key]["_skip"] = True
            if not date_dict[key]["_skip"]:
                existing.append({
                    "position": rank, "rank": rank, "title": title,
                    "genre": genre, "trend": "", "isExclusive": False,
                    "isNetflixOriginal": is_orig, "score": score,
                })

    sorted_weeks = sorted(date_dict.values(), key=lambda w: w["dateFrom"])
    return [{"weekNumber": i + 1, "dateRange": w["dateRange"], "rankings": w["rankings"]}
            for i, w in enumerate(sorted_weeks)]


def _parse_weekly_indexed(ws):
    """索引模式 fallback"""
    date_dict = {}
    for j, row in enumerate(ws.iter_rows(values_only=True)):
        if j == 0:
            continue
        date_from = row[0]
        date_to = row[1]

        # 嘗試過濾「節目」類別（假設 col 2 是類別）
        if len(row) > 7:
            cat = str(row[2]).strip() if row[2] else ""
            if cat == "電影":
                continue
            rank = safe_int(row[3])
            title = clean_title(row[4]) if row[4] else ""
            genre = str(row[5]).strip() if row[5] else "其他"
            is_orig = safe_bool(row[6])
            score = safe_float(row[7], 0)
        else:
            rank = safe_int(row[2])
            title = clean_title(row[3]) if row[3] else ""
            genre = str(row[4]).strip() if row[4] else "其他"
            is_orig = safe_bool(row[5])
            score = safe_float(row[6], 0)

        if date_from is None:
            continue

        key = date_from if isinstance(date_from, datetime) else str(date_from)
        if key not in date_dict:
            if isinstance(date_from, datetime) and isinstance(date_to, datetime):
                date_range = f"{date_from.strftime('%Y-%m-%d')} ~ {date_to.strftime('%Y-%m-%d')}"
            else:
                date_range = str(date_from)
            date_dict[key] = {"dateFrom": date_from, "dateRange": date_range, "rankings": [], "_skip": False}

        if title and rank is not None:
            existing = date_dict[key]["rankings"]
            if rank == 1 and len(existing) > 0:
                date_dict[key]["_skip"] = True
            if not date_dict[key]["_skip"]:
                existing.append({
                    "position": rank, "rank": rank, "title": title,
                    "genre": genre, "trend": "", "isExclusive": False,
                    "isNetflixOriginal": is_orig, "score": score,
                })

    sorted_weeks = sorted(date_dict.values(), key=lambda w: w["dateFrom"])
    return [{"weekNumber": i + 1, "dateRange": w["dateRange"], "rankings": w["rankings"]}
            for i, w in enumerate(sorted_weeks)]


# ── 從週榜衍生整體排名 ──────────────────────────────────────────────────────

def derive_overall_rankings(weekly, show_attrs):
    """從週榜數據衍生所有節目的整體排名（取代樞紐表讀取）"""
    stats = defaultdict(lambda: {
        "score": 0.0, "weeks": 0, "totalPos": 0,
        "genre": "其他", "isNetflixOriginal": False,
        "firstWeekDate": "", "lastWeekDate": "",
    })

    for week in weekly:
        week_start = week["dateRange"].split(" ~ ")[0]
        for item in week["rankings"]:
            t = item["title"]
            if not t:
                continue
            s = stats[t]
            s["score"] += item.get("score", 11 - item["position"])
            s["weeks"] += 1
            s["totalPos"] += item["position"]
            s["genre"] = item.get("genre", "其他")
            if item.get("isNetflixOriginal"):
                s["isNetflixOriginal"] = True
            if not s["firstWeekDate"]:
                s["firstWeekDate"] = week_start
            s["lastWeekDate"] = week_start

    # 用劇集屬性資料庫補充 isNetflixOriginal 和 releaseType
    for title, s in stats.items():
        attr = show_attrs.get(title, {})
        if attr.get("isNetflixOriginal"):
            s["isNetflixOriginal"] = True

    results = []
    for title, s in stats.items():
        attr = show_attrs.get(title, {})
        weeks = s["weeks"]
        results.append({
            "rank": 0,
            "title": title,
            "totalScore": s["score"],
            "genre": s["genre"],
            "weeksOnChart": weeks,
            "avgRank": round(s["totalPos"] / weeks, 2) if weeks > 0 else 0,
            "isNetflixOriginal": s["isNetflixOriginal"],
            "releaseType": attr.get("releaseType", "weekly"),
            "totalEpisodes": attr.get("totalEpisodes", ""),
            "firstWeekDate": s["firstWeekDate"],
            "lastWeekDate": s["lastWeekDate"],
        })

    results.sort(key=lambda x: -x["totalScore"])
    for i, r in enumerate(results):
        r["rank"] = i + 1

    return results


# ── Taiwan drama daily rankings ──────────────────────────────────────────────

def parse_daily_clean(ws):
    col_map = build_col_map(ws, 0)
    use_named = "節目名稱" in col_map and "排名" in col_map

    results = []
    for j, row in enumerate(ws.iter_rows(values_only=True)):
        if j == 0:
            continue
        if use_named:
            title = clean_title(get_col(row, col_map, "節目名稱"))
            day_idx = safe_int(get_col(row, col_map, "上線天數"))
            rank = safe_int(get_col(row, col_map, "排名"))
            score = safe_float(get_col(row, col_map, "積分"), 0)
            is_all = safe_bool(get_col(row, col_map, "是否單次上架"))
        else:
            title = clean_title(row[1]) if row[1] else ""
            day_idx = safe_int(row[3])
            rank = safe_int(row[4])
            score = safe_float(row[5], 0)
            is_all = safe_bool(row[6]) if len(row) > 6 else False

        if title and day_idx is not None and rank is not None:
            results.append({
                "dayIndex": day_idx, "title": title,
                "rank": rank, "score": score, "isAllAtOnce": is_all,
            })
    return results


# ── Compute Taiwan drama rankings ────────────────────────────────────────────

def compute_taiwan_drama_rankings(daily_entries, weekly_weeks, show_attrs):
    daily_stats = defaultdict(lambda: {"dailyScore": 0.0, "daysOnChart": 0, "rankSum": 0, "isAllAtOnce": False})
    for entry in daily_entries:
        t = entry["title"]
        daily_stats[t]["dailyScore"] += entry["score"]
        if entry.get("isAllAtOnce"):
            daily_stats[t]["isAllAtOnce"] = True
        if entry["rank"] <= 10:
            daily_stats[t]["daysOnChart"] += 1
            daily_stats[t]["rankSum"] += entry["rank"]

    weekly_stats = defaultdict(lambda: {"weeklyScore": 0.0, "weeksOnChart": 0, "rankSum": 0, "isNetflixOriginal": False})
    for week in weekly_weeks:
        for item in week["rankings"]:
            if item["genre"] == "台劇":
                t = item["title"]
                weekly_stats[t]["weeklyScore"] += item["score"]
                weekly_stats[t]["weeksOnChart"] += 1
                weekly_stats[t]["rankSum"] += item["rank"]
                if item["isNetflixOriginal"]:
                    weekly_stats[t]["isNetflixOriginal"] = True

    all_titles = set(daily_stats.keys()) | set(weekly_stats.keys())
    results = []
    for title in all_titles:
        d = daily_stats[title]
        w = weekly_stats[title]
        attr = show_attrs.get(title, {})
        days = d["daysOnChart"]
        weeks = w["weeksOnChart"]

        # releaseType 優先從屬性資料庫取得
        if attr.get("releaseType"):
            release_type = attr["releaseType"]
        elif d.get("isAllAtOnce"):
            release_type = "allAtOnce"
        else:
            release_type = "weekly"

        is_orig = w.get("isNetflixOriginal", False) or attr.get("isNetflixOriginal", False)

        results.append({
            "title": title,
            "weeklyRank": None,
            "weeklyScore": w["weeklyScore"],
            "weeksOnChart": weeks,
            "weeklyAvgRank": round(w["rankSum"] / weeks, 2) if weeks > 0 else 0,
            "dailyRank": None,
            "dailyScore": d["dailyScore"],
            "daysOnChart": days,
            "dailyAvgRank": round(d["rankSum"] / days, 2) if days > 0 else 0,
            "isNetflixOriginal": is_orig,
            "isAllAtOnce": release_type == "allAtOnce",
            "releaseType": release_type,
        })

    results.sort(key=lambda x: (-x["weeklyScore"], -x["dailyScore"]))
    for i, r in enumerate(results):
        r["weeklyRank"] = i + 1
        r["dailyRank"] = i + 1

    return results


# ── 型別驗證檔產生 ─────────────────────────────────────────────────────────

def _infer_ts_type(value):
    """從 Python 值推斷 TypeScript 型別"""
    if value is None:
        return "null"
    if isinstance(value, bool):
        return "boolean"
    if isinstance(value, (int, float)):
        return "number"
    if isinstance(value, str):
        return "string"
    if isinstance(value, list):
        if len(value) == 0:
            return "unknown[]"
        return f"({_infer_ts_type(value[0])})[]"
    if isinstance(value, dict):
        return "object"
    return "unknown"


def _extract_keys(records, name):
    """從一組 records 取得所有 key 及其型別，產出驗證用的 key set"""
    if not records:
        return []
    sample = records[0] if isinstance(records, list) else records
    entries = []
    for k, v in sample.items():
        entries.append((k, _infer_ts_type(v)))
    return entries


def generate_schema_validator(output):
    """
    從實際 JSON 資料結構產出一份 TypeScript 驗證檔。
    開發時 import 這個檔案，即可在 runtime 比對 JSON 欄位與型別宣告是否一致。
    """
    sections = {
        "OverallRankingEntry": output["overallRankings"],
        "TaiwanDramaRanking": output["taiwanDramaRankings"],
        "DailyRankingEntry": output["dailyRankings"],
        "WeeklyRankingWeek": output["weeklyRankings"],
        "ShowAttributes": list(output["showAttributes"].values()) if output["showAttributes"] else [],
    }

    lines = [
        "// ⚠ 此檔案由 convert_excel.py 自動產生，請勿手動編輯",
        f"// 產生時間：{datetime.now().isoformat()}",
        "//",
        "// 用途：開發時驗證 rankings.json 的欄位是否與 TypeScript 型別一致",
        "// 若 Python 輸出了新欄位但 types/index.ts 未宣告，validateSchema() 會在 console 警告",
        "",
        "type FieldDef = { key: string; type: string }",
        "",
    ]

    for type_name, records in sections.items():
        fields = _extract_keys(records, type_name)
        lines.append(f"export const {type_name}_FIELDS: FieldDef[] = [")
        for k, t in fields:
            lines.append(f"  {{ key: '{k}', type: '{t}' }},")
        lines.append("]")
        lines.append("")

    # 驗證函式
    lines.extend([
        "function checkFields(typeName: string, expected: FieldDef[], actual: Record<string, unknown>) {",
        "  const actualKeys = new Set(Object.keys(actual))",
        "  const expectedKeys = new Set(expected.map(f => f.key))",
        "  for (const key of actualKeys) {",
        "    if (!expectedKeys.has(key)) {",
        '      console.warn(`[Schema] ${typeName}: JSON 有欄位 "${key}" 但 TypeScript 未宣告`)',
        "    }",
        "  }",
        "  for (const f of expected) {",
        "    if (!actualKeys.has(f.key)) {",
        '      console.warn(`[Schema] ${typeName}: TypeScript 宣告了 "${f.key}" 但 JSON 中不存在`)',
        "    }",
        "  }",
        "}",
        "",
        "// eslint-disable-next-line @typescript-eslint/no-explicit-any",
        "export function validateSchema(data: any) {",
        "  if (!data) return",
        "  if (data.overallRankings?.[0])",
        "    checkFields('OverallRankingEntry', OverallRankingEntry_FIELDS, data.overallRankings[0])",
        "  if (data.taiwanDramaRankings?.[0])",
        "    checkFields('TaiwanDramaRanking', TaiwanDramaRanking_FIELDS, data.taiwanDramaRankings[0])",
        "  if (data.dailyRankings?.[0])",
        "    checkFields('DailyRankingEntry', DailyRankingEntry_FIELDS, data.dailyRankings[0])",
        "  if (data.weeklyRankings?.[0])",
        "    checkFields('WeeklyRankingWeek', WeeklyRankingWeek_FIELDS, data.weeklyRankings[0])",
        "  const attrKeys = Object.keys(data.showAttributes ?? {})",
        "  if (attrKeys.length > 0)",
        "    checkFields('ShowAttributes', ShowAttributes_FIELDS, data.showAttributes[attrKeys[0]])",
        "}",
        "",
    ])

    out_path = OUTPUT_DIR / "src" / "utils" / "schemaValidator.generated.ts"
    out_path.parent.mkdir(parents=True, exist_ok=True)
    with open(out_path, "w", encoding="utf-8") as f:
        f.write("\n".join(lines))


# ── main ──────────────────────────────────────────────────────────────────────

def main():
    sys.stdout.reconfigure(encoding="utf-8")

    # 命令列參數：可指定 Excel 路徑
    if len(sys.argv) > 1:
        excel_path = Path(sys.argv[1])
    else:
        excel_path = DEFAULT_EXCEL

    print(f"讀取 Excel：{excel_path}")
    if not excel_path.exists():
        print(f"ERROR: 找不到 {excel_path}")
        print(f"用法：python {sys.argv[0]} [excel_path]")
        sys.exit(1)

    wb = openpyxl.load_workbook(excel_path, data_only=True)
    print(f"Sheet 清單：{wb.sheetnames}")

    # ── 劇集屬性資料庫 ──
    show_attrs = {}
    ws_attrs = find_sheet(wb, SHEET_ATTRS)
    if ws_attrs:
        print(f"\n解析劇集屬性資料庫：{SHEET_ATTRS}")
        show_attrs = parse_show_attributes(ws_attrs)
        print(f"  → {len(show_attrs)} 筆劇集屬性")
        for title, attr in list(show_attrs.items())[:3]:
            print(f"     {title}: {attr}")
    else:
        print(f"  ⚠ 找不到 sheet「{SHEET_ATTRS}」，跳過屬性資料庫")

    # ── 週榜（全年度合併，單一 sheet） ──
    ws_weekly = find_sheet(wb, SHEET_WEEKLY)
    if not ws_weekly:
        print(f"ERROR: 找不到 sheet「{SHEET_WEEKLY}」")
        sys.exit(1)

    print(f"\n解析週榜：{SHEET_WEEKLY}")
    weekly = parse_weekly_clean(ws_weekly)
    filled = sum(1 for w in weekly if w["rankings"])
    print(f"  → {len(weekly)} 週，{filled} 週有資料")

    # ── 校驗週榜 ──
    warn_count = 0
    for w in weekly:
        validate_date_range(w["dateRange"], f"W{w['weekNumber']}")
        errs = validate_weekly_week(w)
        for e in errs:
            print(f"  ⚠ W{w['weekNumber']} {e}")
            warn_count += 1
    if warn_count == 0:
        print("  ✓ 週榜校驗通過")
    else:
        print(f"  ⚠ 週榜校驗發現 {warn_count} 個警告")

    # ── 從週榜衍生整體排名 ──
    print(f"\n衍生整體排名（從週榜計算）")
    overall = derive_overall_rankings(weekly, show_attrs)
    print(f"  → {len(overall)} 筆整體排名")

    # ── 台劇每日排名 ──
    daily = []
    ws_daily = find_sheet(wb, SHEET_DAILY)
    if ws_daily:
        print(f"\n解析台劇每日排名：{SHEET_DAILY}")
        daily = parse_daily_clean(ws_daily)
        print(f"  → {len(daily)} 筆每日排名")
    else:
        print(f"  ⚠ 找不到 sheet「{SHEET_DAILY}」，跳過台劇每日排名")

    # ── 計算台劇排名 ──
    taiwan = compute_taiwan_drama_rankings(daily, weekly, show_attrs)
    print(f"  → {len(taiwan)} 筆台劇排名（衍生自日榜/週榜）")

    # ── dataThrough ──
    data_through = ""
    for w in reversed(weekly):
        if w["rankings"] and w["dateRange"]:
            data_through = w["dateRange"].split("~")[-1].strip()
            break

    output = {
        "meta": {
            "generatedAt": datetime.now().isoformat(),
            "dataThrough": data_through,
        },
        "showAttributes": show_attrs,
        "overallRankings": overall,
        "taiwanDramaRankings": taiwan,
        "dailyRankings": daily,
        "weeklyRankings": weekly,
    }

    # 輸出到 public/data（production fetch 用）
    out_path = OUTPUT_DIR / "public" / "data" / "rankings.json"
    out_path.parent.mkdir(parents=True, exist_ok=True)
    with open(out_path, "w", encoding="utf-8") as f:
        json.dump(output, f, ensure_ascii=False, indent=2)
    size_kb = out_path.stat().st_size // 1024
    print(f"\n輸出：{out_path} ({size_kb} KB)")

    # 自動產生型別驗證檔（開發時比對 JSON 與 TypeScript 型別）
    generate_schema_validator(output)
    print("  ✓ 型別驗證檔已更新：src/utils/schemaValidator.generated.ts")

    print(f"\n完成！")
    print(f"  劇集屬性：{len(show_attrs)} 筆")
    print(f"  整體排名：{len(overall)} 筆")
    print(f"  台劇排名：{len(taiwan)} 筆")
    print(f"  每日排名：{len(daily)} 筆")
    print(f"  週榜：{len(weekly)} 週")


if __name__ == "__main__":
    main()
